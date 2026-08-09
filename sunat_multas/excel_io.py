from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

from .models import Company, ManifestRecord
from .reportes import TipoReporte


REQUIRED_LIST_COLUMNS = {"ITEM", "NOMBRE", "RUC", "USUARIO", "CLAVE"}
TARGET_SHEETS = ("IMPO118", "IMPO235", "EXPO118", "EXPO235")

MANIFEST_SHEET_COLUMNS = [
    "Manifiesto de Carga",
    "Fecha del Manifiesto de Carga",
    "Manifiesto Desconsolidado",
    "Fecha del Manifiesto Desconsolidado",
    "Agente de Carga RUC",
    "Número de Ticket",
    "Fecha de Llegada",
    "Fecha de Término de la Descarga",
    "Estado del Manifiesto Desconsolidado",
]

MANIFEST_SOURCE_SHEET = "IMPO118-Transmisiones"
MANIFEST_DEST_SHEET = "IMPO118"
DESCONS_SOURCE_COLUMN = 3
DESCONS_DEST_COLUMN = 3
CARGA_SOURCE_COLUMN = 1
CARGA_DEST_COLUMN = 4
LLEGADA_SOURCE_COLUMN = 7
LLEGADA_DEST_COLUMN = 12

DETALLE_COLUMNS = {
    "master": 2,
    "puerto_embarque": 5,
    "cnt": 7,
    "fecha_hijo": 9,
    "fecha_master": 10,
    "fecha_info": 11,
}

TIPO_NUMERACION_COLUMN = 8
PUERTO_COLUMN = 6


def safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', " ", value).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned[:180] or "empresa_sin_nombre"


def read_companies(list_path: Path) -> list[Company]:
    workbook = load_workbook(list_path, read_only=True, data_only=True)
    if "LISTA" not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja LISTA en {list_path}")

    sheet = workbook["LISTA"]
    header_values = [str(cell.value).strip().upper() if cell.value is not None else "" for cell in sheet[1]]
    header_index = {name: index + 1 for index, name in enumerate(header_values) if name}
    missing = REQUIRED_LIST_COLUMNS - set(header_index)
    if missing:
        raise ValueError(f"Faltan columnas en LISTA: {', '.join(sorted(missing))}")

    companies: list[Company] = []
    seen_keys: set[tuple[str, str]] = set()
    for row_number in range(2, sheet.max_row + 1):
        name = _cell_text(sheet.cell(row_number, header_index["NOMBRE"]).value)
        ruc = _cell_text(sheet.cell(row_number, header_index["RUC"]).value)
        user = _cell_text(sheet.cell(row_number, header_index["USUARIO"]).value)
        password = _cell_text(sheet.cell(row_number, header_index["CLAVE"]).value)
        item = _cell_text(sheet.cell(row_number, header_index["ITEM"]).value)
        if not any([name, ruc, user, password]):
            continue
        if not all([name, ruc, user, password]):
            raise ValueError(f"Fila {row_number} incompleta en LISTA")
        key = (ruc, user)
        if key in seen_keys:
            raise ValueError(f"Empresa duplicada en LISTA: RUC {ruc}, usuario {user}")
        seen_keys.add(key)
        companies.append(Company(item=item, name=name, ruc=ruc, user=user, password=password))

    return companies


def create_company_workbook(
    template_path: Path,
    output_dir: Path,
    company: Company,
    keep_existing: bool,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{safe_filename(company.name)}.xlsx"
    if output_path.exists() and keep_existing:
        return output_path
    shutil.copy2(template_path, output_path)
    return output_path


def append_records(workbook_path: Path, records: list[ManifestRecord]) -> int:
    workbook = load_workbook(workbook_path)
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"La plantilla no contiene la hoja {sheet_name}")

    inserted = 0
    counters = {sheet_name: _next_sequence(workbook[sheet_name]) for sheet_name in TARGET_SHEETS}
    for record in records:
        sheet_name = classify_sheet(record)
        sheet = workbook[sheet_name]
        row = sheet.max_row + 1
        sequence = counters[sheet_name]
        counters[sheet_name] += 1

        if sheet_name.startswith("IMPO"):
            values = [
                sequence,
                record.master,
                record.process_number,
                record.manifest_number,
                record.port_code,
                record.port,
                record.cnt,
                record.numbering_type,
                record.initial_numbering_datetime,
                record.line_transmission_datetime,
                record.complementary_info_datetime,
                record.vessel_arrival_datetime,
                record.fine_status,
            ]
        else:
            values = [
                sequence,
                record.master,
                record.process_number,
                record.manifest_number,
                record.port_code,
                record.port,
                record.cargo_agent_transmission_datetime,
                record.boarding_end_datetime,
                record.fine_status,
            ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        inserted += 1

    workbook.save(workbook_path)
    return inserted


def _crear_hoja_transmisiones(workbook, rows: list[list[str]], tipo: TipoReporte) -> None:
    if tipo.hoja_destino not in workbook.sheetnames:
        raise ValueError(f"La plantilla no contiene la hoja {tipo.hoja_destino}")
    if tipo.hoja_transmisiones in workbook.sheetnames:
        del workbook[tipo.hoja_transmisiones]
    sheet = workbook.create_sheet(title=tipo.hoja_transmisiones)
    sheet.append(list(tipo.cabeceras_transmisiones))
    for row in rows:
        sheet.append((row + [None] * tipo.columnas_extraccion)[:tipo.columnas_extraccion])


def escribir_hoja_transmisiones(workbook_path: Path, rows: list[list[str]], tipo: TipoReporte) -> int:
    workbook = load_workbook(workbook_path)
    _crear_hoja_transmisiones(workbook, rows, tipo)
    workbook.save(workbook_path)
    return len(rows)


def append_manifest_sheet(workbook_path: Path, rows: list[list[str]], tipo: TipoReporte) -> int:
    workbook = load_workbook(workbook_path)
    _crear_hoja_transmisiones(workbook, rows, tipo)
    copy_manifiestos_to_impo118(workbook, tipo)
    workbook.save(workbook_path)
    return len(rows)


def read_transmisiones(workbook_path: Path, tipo: TipoReporte, source_title: str | None = None) -> list[str]:
    workbook = load_workbook(workbook_path, data_only=True)
    source_title = source_title or tipo.hoja_transmisiones
    if source_title not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {source_title!r} en {workbook_path}")
    sheet = workbook[source_title]
    codes: list[str] = []
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, DESCONS_SOURCE_COLUMN).value
        codes.append(str(value).strip() if value is not None else "")
    return codes


def read_proceso_destino(workbook_path: Path, tipo: TipoReporte) -> list[str]:
    workbook = load_workbook(workbook_path, data_only=True)
    if tipo.hoja_destino not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {tipo.hoja_destino!r} en {workbook_path}")
    sheet = workbook[tipo.hoja_destino]
    values: list[str] = []
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 3).value
        text = str(value).strip() if value is not None else ""
        if text:
            values.append(text)
    return values


def leer_mapa_master_fila(workbook_path: Path, tipo: TipoReporte) -> dict[str, int]:
    """Mapea cada MASTER (columna B) a su fila en la hoja destino."""
    workbook = load_workbook(workbook_path, data_only=True)
    if tipo.hoja_destino not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {tipo.hoja_destino!r} en {workbook_path}")
    sheet = workbook[tipo.hoja_destino]
    mapa: dict[str, int] = {}
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        text = str(value).strip() if value is not None else ""
        if text:
            mapa[text] = row
    return mapa


def read_proceso_destino_grupos(workbook_path: Path, tipo: TipoReporte) -> list[dict]:
    """Devuelve grupos de códigos consecutivos con su fila de inicio en la hoja destino."""
    workbook = load_workbook(workbook_path, data_only=True)
    if tipo.hoja_destino not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {tipo.hoja_destino!r} en {workbook_path}")
    sheet = workbook[tipo.hoja_destino]
    grupos: list[dict] = []
    fila_inicio = None
    codigo_actual = None
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 3).value
        text = str(value).strip() if value is not None else ""
        if not text:
            continue
        if text != codigo_actual:
            codigo_actual = text
            fila_inicio = row
            grupos.append({"code": codigo_actual, "start_row": fila_inicio, "count": 1})
        else:
            grupos[-1]["count"] += 1
    return grupos


def escribir_puerto_destino(workbook_path: Path, tipo: TipoReporte, fila: int, texto_puerto: str) -> None:
    codigo_puerto, _, puerto = texto_puerto.partition("-")
    workbook = load_workbook(workbook_path)
    sheet = workbook[tipo.hoja_destino]
    sheet.cell(fila, 5, codigo_puerto.strip() or None)
    sheet.cell(fila, 6, puerto.strip() or None)
    workbook.save(workbook_path)


def _tiene_fecha(valor: object) -> bool:
    if valor is None:
        return False
    if isinstance(valor, (int, float)):
        return valor != 0
    return bool(str(valor).strip())


def cargar_mapa_puertos(ruta_json: Path) -> dict[str, str]:
    """Carga el JSON y crea un diccionario con clave de 5 caracteres: pais + codigo_actual."""
    mapa: dict[str, str] = {}
    with ruta_json.open("r", encoding="utf-8") as json_file:
        datos = json.load(json_file)
    for item in datos:
        pais = str(item.get("pais", "") or "").strip()
        codigo_local = str(item.get("codigo_actual", "") or "").strip()
        descripcion = str(item.get("descripcion", "") or "").strip()
        codigo_5_letras = f"{pais}{codigo_local}"
        if codigo_5_letras and codigo_5_letras not in mapa:
            mapa[codigo_5_letras] = descripcion
    return mapa


def aplicar_detalle_filas(
    workbook,
    updates: list[tuple[int, dict]],
    tipo: TipoReporte,
    mapa_puertos: dict[str, str] | None = None,
) -> None:
    sheet = workbook[tipo.hoja_destino]
    columnas = dict(tipo.columnas_detalle)
    for fila, datos in updates:
        for key, column in columnas.items():
            value = datos.get(key)
            if value:
                sheet.cell(fila, column, value)
        if mapa_puertos:
            codigo_puerto = sheet.cell(fila, columnas["puerto_embarque"]).value
            if codigo_puerto:
                codigo = str(codigo_puerto).strip()
                if codigo in mapa_puertos:
                    sheet.cell(fila, PUERTO_COLUMN, mapa_puertos[codigo])
        fechas = [
            sheet.cell(fila, columnas[key]).value
            for key in ("fecha_hijo", "fecha_master", "fecha_info")
        ]
        tipo_num = "PREVIO" if all(_tiene_fecha(fecha) for fecha in fechas) else ""
        sheet.cell(fila, tipo.columna_tipo_numeracion, tipo_num)


def load_procesados(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return set(data.get("procesados", []))
    except (json.JSONDecodeError, OSError):
        return set()


def save_procesados(path: Path, codes: set[str]) -> None:
    path.write_text(
        json.dumps({"procesados": sorted(codes)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def copy_manifiestos_to_impo118(workbook, tipo: TipoReporte, source_title: str | None = None) -> int:
    source = workbook[source_title or tipo.hoja_transmisiones]
    destination = workbook[tipo.hoja_destino]
    copied = 0

    mappings = tipo.mapeo_copia
    for src_col, dst_col, transform in mappings:
        values: list[object] = []
        for row in range(2, source.max_row + 1):
            raw = source.cell(row, src_col).value
            if raw is None or not str(raw).strip():
                continue
            if transform:
                segment = _ultimo_segmento(str(raw))
                if not segment:
                    continue
                values.append(segment)
            else:
                values.append(raw)
        _clear_column(destination, dst_col)
        for index, value in enumerate(values):
            destination.cell(2 + index, dst_col, value)
        copied += len(values)

    _number_destino_rows(destination, tipo)
    if tipo.columnas_condicion_color:
        _aplicar_color_condicion(source, destination, tipo)
    return copied


def _aplicar_color_condicion(source, destination, tipo: TipoReporte) -> None:
    from openpyxl.styles import PatternFill

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    col1, col2 = tipo.columnas_condicion_color
    dst_col = tipo.columna_color_destino
    for row in range(2, source.max_row + 1):
        value1 = source.cell(row, col1).value
        value2 = source.cell(row, col2).value
        if value1 is None and value2 is None:
            continue
        solo_guiones1 = _es_solo_guiones(value1)
        solo_guiones2 = _es_solo_guiones(value2)
        cell = destination.cell(row, dst_col)
        if solo_guiones1 and solo_guiones2:
            cell.fill = verde
        elif _tiene_texto_no_guiones(value1) or _tiene_texto_no_guiones(value2):
            cell.fill = rojo


def _es_solo_guiones(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text != "" and all(ch == "-" for ch in text)


def _tiene_texto_no_guiones(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text != "" and any(ch != "-" for ch in text)


def _number_destino_rows(destination, tipo: TipoReporte) -> None:
    data_columns = tuple(dst_col for _, dst_col, _ in tipo.mapeo_copia)
    count = 0
    for row in range(2, destination.max_row + 1):
        if any(destination.cell(row, col).value not in (None, "") for col in data_columns):
            count += 1
        else:
            break
    _clear_column(destination, 1)
    for index in range(count):
        destination.cell(2 + index, 1, index + 1)


def _ultimo_segmento(value: str) -> str:
    text = value.strip()
    if "-" in text:
        return text.rsplit("-", 1)[-1].strip()
    return text


def _clear_column(sheet, column: int) -> None:
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, column).value = None


def process_manifiestos_excel(
    workbook_path: Path,
    tipo: TipoReporte,
    source_title: str | None = None,
) -> int:
    workbook = load_workbook(workbook_path)
    source_title = source_title or tipo.hoja_transmisiones
    if source_title not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {source_title!r} en {workbook_path}")
    if tipo.hoja_destino not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {tipo.hoja_destino} en {workbook_path}")
    copied = copy_manifiestos_to_impo118(workbook, tipo, source_title)
    workbook.save(workbook_path)
    return copied


def classify_sheet(record: ManifestRecord) -> str:
    operation = _normalize(record.operation)
    manifest_type = _normalize(record.manifest_type)
    if "EXPO" in operation or "EXPORT" in operation or "SALIDA" in operation:
        prefix = "EXPO"
    else:
        prefix = "IMPO"

    if "235" in manifest_type:
        suffix = "235"
    elif "118" in manifest_type:
        suffix = "118"
    else:
        raise ValueError(f"No se pudo clasificar manifiesto: {record.raw or record}")

    return f"{prefix}{suffix}"


def _next_sequence(sheet) -> int:
    for row in range(sheet.max_row, 1, -1):
        value = sheet.cell(row, 1).value
        if isinstance(value, int):
            return value + 1
        if isinstance(value, str) and value.strip().isdigit():
            return int(value.strip()) + 1
    return 1


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize(value: str) -> str:
    return (value or "").strip().upper()
